import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
import time
import pandas as pd
from openpyxl import load_workbook

DOI = []
TITLE = []
NUM_AUTHORS = []
NUM_AFFILIATIONS = []
AUTHORS = []
YEAR = []
JOURNAL = []
DOCUMENT_TYPE = []
COUNTRIES = []
SEARCH = []
DATES = []

chrome_driver_path = "C:\Development\chromedriver.exe"
driver = webdriver.Chrome(executable_path=chrome_driver_path)

driver.get("https://www.scopus.com/search/form.uri?display=basic#basic")
# searching details
search = input("Search documents: \n")
SEARCH.append(search)
date = input("Do you want to specify dates?(Yes/No)")
if date.capitalize() == "Yes":
    driver.find_element(By.CLASS_NAME, 'flex-grow-1').send_keys(search)
    driver.find_element(By.XPATH,
                        "/html/body/div/div/div[1]/div[2]/div/div[3]/div/div[2]/div["
                        "2]/micro-ui/scopus-homepage/div/div/els-tab/els-tab-panel[1]/div/form/div[2]/div[1]/button["
                        "2]/span[2]").click()
    time.sleep(1)
    starting_date = input("Put starting year.")
    to_date = input("Put end date.")
    DATES.append(starting_date)
    DATES.append(to_date)
    drop_menu_from = Select(driver.find_element(By.XPATH,
                                                "/html/body/div/div/div[1]/div[2]/div/div[3]/div/div[2]/div["
                                                "2]/micro-ui/scopus-homepage/div/div/els-tab/els-tab-panel["
                                                "1]/div/form/div[2]/div[1]/els-select/div/label/select"))
    drop_menu_from.select_by_visible_text(starting_date)
    drop_menu_to = Select(driver.find_element(By.XPATH,
                                              "/html/body/div/div/div[1]/div[2]/div/div[3]/div/div[2]/div["
                                              "2]/micro-ui/scopus-homepage/div/div/els-tab/els-tab-panel["
                                              "1]/div/form/div[2]/div[2]/els-select/div/label/select"))
    drop_menu_to.select_by_visible_text(to_date)
    driver.find_element(By.XPATH,
                        '/html/body/div/div/div[1]/div[2]/div/div[3]/div/div[2]/div['
                        '2]/micro-ui/scopus-homepage/div/div/els-tab/els-tab-panel[1]/div/form/div[4]/div['
                        '2]/button/span[1]').click()
else:
    DATES = ["XXX", "YYY"]
    driver.find_element(By.CLASS_NAME, 'flex-grow-1').send_keys(search)
    driver.find_element(By.XPATH,
                        "/html/body/div/div/div[1]/div[2]/div/div[3]/div/div[2]/div["
                        "2]/micro-ui/scopus-homepage/div/div/els-tab/els-tab-panel[1]/div/form/div[2]/div["
                        "2]/button").click()
time.sleep(2)
doc_num = int(driver.find_element(By.XPATH,
                                  "/html/body/div[1]/div/div[1]/div/div/div[3]/form/div[1]/div/header/h1/span[1]").text.replace(
    ",", ""))

# checking the nuber of pubs (if more than 200 better to refine the search)
if doc_num < 2000:
    print(f"There is {doc_num} documents.")
elif doc_num > 2000:
    print(f"There is {doc_num} documents. Try to refine your search to be under 2000.")

counter = 0
refining_on = True
while refining_on:

    refine = input("Do you want to refine your search?(Yes/No)")

    if refine.capitalize() == "Yes":
        search_ref = input("Search documents: \n")
        SEARCH.append(search_ref)
        driver.find_element(By.XPATH,
                            "/html/body/div[1]/div/div[1]/div/div/div[3]/form/div[4]/div[1]/div/div/div/div["
                            "3]/div/div[1]/div/div/span/input").send_keys(
            search_ref)
        driver.find_element(By.XPATH,
                            '/html/body/div[1]/div/div[1]/div/div/div[3]/form/div[4]/div[1]/div/div/div/div['
                            '3]/div/div[1]/div/div/span/button[2]').click()
        doc_num = int(driver.find_element(By.XPATH,
                                          "/html/body/div[1]/div/div[1]/div/div/div[3]/form/div["
                                          "1]/div/header/h1/span[1]").text.replace(",", ""))
    elif refine.capitalize() == "No":
        refining_on = False

searching_is_on = True

while searching_is_on:
    start = time.process_time()
    num_per_page = len(driver.find_elements(By.CLASS_NAME, "searchArea"))
    for n in range(0, num_per_page):
        time.sleep(0.5)
# finding type of paper
        if doc_num < 2000:
            try:
                year = driver.find_element(By.ID, f"resultDataRow{n}").find_elements(By.TAG_NAME, "td")[2].text
                YEAR.append(year)
                driver.find_element(By.ID, f"resultDataRow{n}").find_element(By.TAG_NAME, "a").click()
                counter = counter + 1
                # print(counter)
                # print(doc_num)
                # print(num_per_page)
                time.sleep(10)
                art_type = driver.find_element(By.XPATH,
                                               "/html/body/div/div/div[1]/div[2]/div/div[3]/div[3]/div/div[1]/div["
                                               "2]/micro-ui/scopus-document-details-page/div/article/div["
                                               "2]/aside/div/div/div/dl[1]/dd")
                DOCUMENT_TYPE.append(art_type.text)
            except:
                DOCUMENT_TYPE.append("Not given - ŁZ")
# finding DOI
            try:
                doi = driver.find_element(By.XPATH,
                                          "/html/body/div/div/div[1]/div[2]/div/div[3]/div[3]/div/div[1]/div["
                                          "2]/micro-ui/scopus-document-details-page/div/article/div["
                                          "2]/aside/div/div/div/dl[4]/dd")
                DOI.append(doi.text)
            except:
                DOI.append("Not given - ŁZ")
# finding titles
            try:
                title = driver.find_element(By.XPATH,
                                            "/html/body/div/div/div[1]/div[2]/div/div[3]/div[3]/div/div[1]/div["
                                            "2]/micro-ui/scopus-document-details-page/div/article/div[2]/section/div["
                                            "1]/div/els-typography/span")
                print(title.text)
                TITLE.append(title.text)
            except:
                TITLE.append("Not given - ŁZ")
 # finding num of authors
            try:
                num_authors = len(
                    driver.find_element(By.XPATH, "/html/body/div/div/div[1]/div[2]/div/div[3]/div[3]/div/div["
                                                  "1]/div["
                                                  "2]/micro-ui/scopus-document-details-page/div/article/div["
                                                  "2]/section/div[2]/div/ul").find_elements(By.TAG_NAME, "li"))
                NUM_AUTHORS.append(num_authors)
            except:
                NUM_AUTHORS.append("Not given - ŁZ")
# num of affilations
            try:
                num_affil = len(
                    driver.find_element(By.XPATH, "/html/body/div/div/div[1]/div[2]/div/div[3]/div[3]/div/div[1]/div["
                                                  "2]/micro-ui/scopus-document-details-page/div/article/div["
                                                  "2]/section/div[3]/div/div/ul").find_elements(By.TAG_NAME, "li"))
                NUM_AFFILIATIONS.append(num_affil)
            except:
                NUM_AFFILIATIONS.append("Not given - ŁZ")
# countires
            try:
                country_list = []
                countries = driver.find_element(By.XPATH,
                                                "/html/body/div/div/div[1]/div[2]/div/div[3]/div[3]/div/div[1]/div["
                                                "2]/micro-ui/scopus-document-details-page/div/article/div["
                                                "2]/section/div[3]/div/div/ul").find_elements(By.TAG_NAME, "span")
                for i in range(0, len(countries)):
                    country = countries[i].text.split(",")[-1]
                    if country not in country_list:
                        country_list.append(country)
                    else:
                        continue
                COUNTRIES.append(country_list)
            except:
                COUNTRIES.append(["Not given - ŁZ"])
# finding a journal
            try:
                journal = driver.find_element(By.XPATH,
                                              '/html/body/div/div/div[1]/div[2]/div/div[3]/div[3]/div/div[1]/div['
                                              '2]/micro-ui/scopus-document-details-page/div/article/div['
                                              '1]/div/div/els-button/els-typography').text
                JOURNAL.append(journal)
            except:
                JOURNAL.append("Not given - ŁZ")
# authors list
            try:
                aut_list = driver.find_element(By.XPATH,
                                               "/html/body/div/div/div[1]/div[2]/div/div[3]/div[3]/div/div[1]/div["
                                               "2]/micro-ui/scopus-document-details-page/div/article/div["
                                               "2]/section/div[2]/div").find_elements(By.TAG_NAME, "li")
                authors = []
                for aut in aut_list:
                    authors.append(aut.text)
                # print(authors)
                AUTHORS.append(authors)
            except:
                AUTHORS.append("Not given - ŁZ")
# which page
            pages = driver.find_element(By.CLASS_NAME,
                                            "recordPageCount").text
            now = pages.split('of')[0]
            last = pages.split('of')[1]
# edit page number if comma in it
            if "," in now:
                now = now.split(",")
                now = ''.join(now)
            if "," in last:
                last = last.split(",")
                last = ''.join(last)
            now = int(now)
            last = int(last)
            if now != last:
                driver.back()
            elif now == last:
                driver.back()
                break
        else:
            print("To much results")
            quit()

    time.sleep(1)
    if now == last:
        searching_is_on = False
        break
    elif 20 <= last <= 40 and now == 20:
        driver.find_element(By.XPATH, "/html/body/div[1]/div/div[1]/div/div/div[3]/form/div[4]/div["
                                      "2]/div/div/section[1]/div/div[4]/div/div[2]/ul/li[2]/a").click()
        time.sleep(2)
        num_per_page = len(driver.find_elements(By.CLASS_NAME, "searchArea"))
    elif last > 40:
        driver.find_element(By.XPATH, "//div[@id='resultsFooter']//a[@title='Next page']").click()
        time.sleep(2)
        num_per_page = len(driver.find_elements(By.CLASS_NAME, "searchArea"))


# authors cleaning
authors_list = []
aut = []
for i in AUTHORS:
    for a in i:
        if "\n" in a:
            a = a.split("\n")[0]
        if "," in a:
            a = a.split(",")
            a = [a[0], a[1]]
            a = ' '.join(a)
        if ";" in a:
            a = a.split(";")
            a = a[0]
        a = a[:-1]
        aut.append(a)
    authors_list.append(aut)
    aut = []
AUTHORS = authors_list

# document type cleaning
doc_types = []
for i in DOCUMENT_TYPE:
    if "•" in i:
        i = i.split("•")[0]
    doc_types.append(i)
DOCUMENT_TYPE = doc_types

# cleaning dates
if len(DATES) == 0:
    DATES == ["Not specified", "Not specified"]

# count years
COUNTED_YEARS = {}
for i in YEAR:
    COUNTED_YEARS[i] = [YEAR.count(i)]

# count Countires
COUNTED_COUNTIRES = {}
list_count = [item for i in COUNTRIES for item in i]
for i in list_count:
    COUNTED_COUNTIRES[i] = [list_count.count(i)]

# count Journals
COUNTED_JOURNALS = {}
for i in JOURNAL:
    COUNTED_JOURNALS[i] = [JOURNAL.count(i)]

# Subject area
try:
    driver.find_element(By.ID, "viewMoreLink_SUBJAREA").click()
    driver.find_element(By.ID, 'viewAllLink_SUBJAREA').click()
    time.sleep(1)
    subject_area_body = driver.find_element(By.ID, "overlayBody_SUBJAREA").find_elements(By.CLASS_NAME, "btnText")
    area = []
    number = []
    for i in subject_area_body:
        if subject_area_body.index(i)%2 == 0:
            area.append(i.text)
        elif subject_area_body.index(i)%2 == 1:
            number.append(i.text)

    COUNTED_AREAS = [dict(zip(area, number))]
except:
    subject_area_body = driver.find_element(By.ID, "clusterAttribute_SUBJAREA").find_elements(By.CLASS_NAME, "btnText")
    area = []
    number = []
    for i in subject_area_body:
        if subject_area_body.index(i) % 2 == 0:
            area.append(i.text)
        elif subject_area_body.index(i) % 2 == 1:
            number.append(i.text)

    COUNTED_AREAS = [dict(zip(area, number))]

RESULTS_DICT = {'Search terms': SEARCH,
                'Dates range': f'{DATES[0]}-{DATES[1]}',
                'No.Pub': doc_num,
                'Title': TITLE,
                'Doi': DOI,
                'No.Authors': NUM_AUTHORS,
                'Authors': AUTHORS,
                'No.Affiliations': NUM_AFFILIATIONS,
                'Countires': COUNTRIES,
                'Year': YEAR,
                'Journal': JOURNAL,
                'Document Type': DOCUMENT_TYPE
                }

MY_RESULTS_DICT = dict([(k, pd.Series(v)) for k, v in RESULTS_DICT.items()])

# general DF
df_scopus_scrap = pd.DataFrame(MY_RESULTS_DICT)

# Counted Years DF
df_counted_years = pd.DataFrame(COUNTED_YEARS)

# Counted Countries DF
df_counted_countries = pd.DataFrame(COUNTED_COUNTIRES)

# Counted Journals DF
df_counted_journals = pd.DataFrame(COUNTED_JOURNALS)

# Counted reas DF
df_counted_areas = pd.DataFrame(COUNTED_AREAS)

print(time.process_time() - start)

# Saving to excel
name = input("Put a name for your file: \n")
path = f"C:\\Users\\Martyna\\PycharmProjects\\pythonProject\\day52\\{name}.xlsx"
wb = openpyxl.Workbook()
wb.save(path)
book = load_workbook(path)
writer = pd.ExcelWriter(path, engine='openpyxl')
writer.book = book

df_scopus_scrap.to_excel(writer, sheet_name="General")
df_counted_years.to_excel(writer, sheet_name="Counted Years")
df_counted_countries.to_excel(writer, sheet_name="Counted Countries")
df_counted_journals.to_excel(writer, sheet_name="Counted Journals")
df_counted_areas.to_excel(writer, sheet_name='Counted Areas')
writer.save()
writer.close()
